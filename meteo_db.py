import sqlite3
from tkinter import messagebox
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import datetime
from datetime import datetime


def create_db_luna():  # Создание базы данных

    db = sqlite3.connect(f'C:/Apps/DataBase/new_luna.db')
    c = db.cursor()

    c.execute("""CREATE TABLE meteo (
                local_date_for_db TEXT, 
                local_time_for_db TEXT,
                date_utc TEXT, 
                time_utc TEXT, 
                wind_direction INTEGER, 
                wind_speed INTEGER, 
                wind_gust INTEGER, 
                visibility INTEGER, 
                weather_condition TEXT, 
                temperature REAL,
                dew_point REAL, 
                humidity INTEGER, 
                qt_clouds INTEGER, 
                qt_lower_clouds INTEGER,
                cloud_base INTEGER, 
                clouds_type TEXT, 
                pressure_heli REAL, 
                pressure_sea_level REAL, 
                wave INTEGER,
                comments TEXT, 
                metar_cod TEXT
               
              )""")
    db.close()


def delete_all_data():  # Удаление всех данных из базы
    global db_dir_name
    db = sqlite3.connect(f'{db_dir_name}luna.db')
    c = db.cursor()
    c.execute(" DELETE FROM meteo ")
    db.commit()
    db.close()


def check_data(date, time, db_dir_name, database_name):
    # Creating a connection to the database
    conn = sqlite3.connect(f'{db_dir_name}{database_name}')
    # Creating a cursor object
    cursor = conn.cursor()

    # Selecting data by time and date
    cursor.execute("SELECT * FROM meteo WHERE time_utc = ? and date_utc = ?", (time, date))

    # Fetching the results
    results = cursor.fetchall()

    # Closing the connection
    conn.close()

    # If data already exists, return True, else return False
    if len(results) > 0:
        return True
    else:
        return False


def insert_data(local_date_for_db, local_time_for_db, date_utc, time_utc, wind_direction, wind_speed, wind_gust,
                visibility, weather_condition, temperature, dew_point, humidity, qt_clouds, qt_lower_clouds,
                cloud_base, clouds_type, pressure_heli, pressure_sea_level, wave, comments, metar_cod, db_dir_name,
                database_name):
    # Checking if data already exists for the specified time and date
    if check_data(date_utc, time_utc, db_dir_name, database_name):
        # Displaying a messagebox asking the user if they want,  to overwrite the data
        result = messagebox.askyesno("Данные уже существуют",
                                     "Информация на выбранную дату и время уже существует в базе данных и будет перезаписана.")
        # If the user chooses to overwrite the data, proceed with the insertion
        if result:
            # Creating a connection to the database
            conn = sqlite3.connect(f'{db_dir_name}{database_name}')
            # Creating a cursor object
            cursor = conn.cursor()

            # Inserting data into the table
            cursor.execute("""UPDATE meteo SET local_date_for_db = ?, local_time_for_db = ?, wind_direction = ?, wind_speed = ?, wind_gust = ?, visibility = ?, 
                                                weather_condition = ?, temperature = ?, dew_point = ?, humidity = ?, 
                                                qt_clouds = ?, qt_lower_clouds = ?, cloud_base = ?, clouds_type = ?, pressure_heli = ?, 
                                                pressure_sea_level = ?, wave = ?, comments = ?, metar_cod = ? WHERE time_utc = ? and date_utc = ?""",
                           (local_date_for_db, local_time_for_db, wind_direction, wind_speed, wind_gust, visibility,
                            weather_condition,
                            temperature, dew_point, humidity, qt_clouds, qt_lower_clouds, cloud_base, clouds_type,
                            pressure_heli, pressure_sea_level, wave, comments, metar_cod, time_utc, date_utc))
            # Committing the changes
            conn.commit()

            # Closing the connection
            conn.close()

            # Displaying a messagebox to confirm the insertion
            messagebox.showinfo("Сохранение данных", "Данные успешно сохранены!")

    else:
        # Creating a connection to the database
        conn = sqlite3.connect(f'{db_dir_name}{database_name}')

        # Creating a cursor object
        cursor = conn.cursor()

        # Inserting data into the table
        cursor.execute("INSERT INTO meteo VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?, ?, ?, ?, ?, ?, ?)",
                       (local_date_for_db, local_time_for_db, date_utc, time_utc, wind_direction, wind_speed, wind_gust,
                        visibility, weather_condition,
                        temperature, dew_point, humidity,
                        qt_clouds, qt_lower_clouds, cloud_base, clouds_type, pressure_heli, pressure_sea_level, wave,
                        comments, metar_cod))

        # Committing the changes
        conn.commit()

        # Closing the connection
        conn.close()

        # Displaying a messagebox to confirm the insertion
        messagebox.showinfo("Информация внесена в базу данных", "Данные успешно сохранены!")


def select_from_db(from_date, to_date, db_dir_name, database_name):
    from_db = str(from_date)
    to_db = str(to_date)

    db = sqlite3.connect(f'{db_dir_name}{database_name}')
    c = db.cursor()
    query = "SELECT date_utc, time_utc, wind_direction, wind_speed, wind_gust, visibility, weather_condition," \
            "temperature,dew_point, humidity, qt_clouds, qt_lower_clouds, cloud_base, clouds_type, pressure_heli, " \
            "pressure_sea_level, wave FROM meteo WHERE local_date_for_db BETWEEN ? AND ?"
    c.execute(query, (from_db, to_db))
    items = c.fetchall()
    db.commit()
    db.close()
    return items


def select_from_db_test():
    global db_dir_name
    db = sqlite3.connect(f'{db_dir_name}luna.db')
    c = db.cursor()

    c.execute("SELECT * FROM meteo")
    items = c.fetchall()

    db.commit()
    db.close()
    return items


def insert_additional_data(data):
    """Функция добавляет данные в базу данных"""
    global db_dir_name
    conn = sqlite3.connect(f'{db_dir_name}luna.db')
    data.to_sql('meteo', conn, if_exists='append', index=False)
    conn.close()


def data_for_excel(from_date, to_date, comments, db_dir_name, database_name, folder_path, folder_name):
    excel_folder_name = folder_path
    excel_folder_path = folder_name
    start_date = from_date
    end_date = to_date
    conn = sqlite3.connect(f'{db_dir_name}{database_name}')
    # Формируем параметризованный запрос
    query = "SELECT local_date_for_db, date_utc, time_utc, wind_direction, wind_speed, wind_gust, visibility, " \
            "weather_condition, temperature, dew_point, humidity, qt_clouds, qt_lower_clouds, " \
            "cloud_base, clouds_type, pressure_heli, pressure_sea_level, wave, comments " \
            "FROM meteo " \
            "WHERE local_date_for_db " \
            "BETWEEN :start_date AND :end_date"

    # Выполняем запрос с использованием параметров
    df = pd.read_sql_query(query, params={"start_date": start_date, "end_date": end_date}, con=conn)

    dic_col = {'date_utc': 'Дата', 'time_utc': 'Время(СГВ)', 'wind_direction': 'Направление ветра(град)',
               'wind_speed': 'Средняя скорость ветра(м/с)', 'wind_gust': 'Максимальный порыв ветра(м/с)',
               'visibility': 'Горизонтальная видимость(км)',
               'weather_condition': 'Атмосферные явления', 'temperature': 'Температура воздуха(град)',
               'dew_point': 'Температура точки росы(град)', 'humidity': 'Влажность воздуха(%)',
               'qt_clouds': 'Общее количество облаков(октанты)',
               'qt_lower_clouds': 'Количество нижнего яруса(октанты)', 'cloud_base': 'Высота НГО(м)',
               'clouds_type': 'Форма облачности', 'pressure_heli': 'Давление на уровне вертолётной площадки(мм.рт.ст.)',
               'pressure_sea_level': 'Давление на уровне моря(гПа)', 'wave': 'Высота преобладающих волн(м)',
               'comments': 'Примечание'}
    df = df.rename(columns=dic_col)
    conn.close()
    df['Горизонтальная видимость(км)'] = df['Горизонтальная видимость(км)'] / 1000
    df['Название платформы'] = "LUN - A"
    df = df[['local_date_for_db', 'Название платформы', 'Дата', 'Время(СГВ)', 'Направление ветра(град)', 'Средняя скорость ветра(м/с)',
             'Максимальный порыв ветра(м/с)', 'Горизонтальная видимость(км)', 'Общее количество облаков(октанты)',
             'Количество нижнего яруса(октанты)', 'Высота НГО(м)', 'Форма облачности', 'Атмосферные явления',
             'Температура воздуха(град)', 'Температура точки росы(град)', 'Влажность воздуха(%)',
             'Давление на уровне вертолётной площадки(мм.рт.ст.)', 'Давление на уровне моря(гПа)',
             'Высота преобладающих волн(м)', 'Примечание']]
    if comments == 0:
        df = df.drop('Примечание', axis=1)

    export_to_excel(df, start_date, end_date, excel_folder_path, excel_folder_name)


def export_to_excel(df, from_date, to_date, folder_path, folder_name):
    dir_name = f'{folder_name}{folder_path}'
    start_date = from_date
    end_date = to_date
    # df['local_date_for_db'] = pd.to_datetime(df['local_date_for_db'])
    # months = df['local_date_for_db'].dt.month.unique()
    # year = df['local_date_for_db'].dt.year.unique()
    # print(months, year)
    df = df.drop('local_date_for_db', axis=1)
    # Преобразование формата даты для выгрузки в Excel
    # Иначе при открытии файла дата не распознаётся как дата
    df['Дата'] = df['Дата'].apply(lambda date: datetime.strptime(date, '%Y-%m-%d').date())

    # Save DataFrame to Excel file
    file_name = f"LUNA_выгрузка данных за период {start_date} - {end_date}.xlsx"
    df.to_excel(f"{dir_name}{file_name}", engine='openpyxl', index=False)

    # Load the Excel file with openpyxl
    book = openpyxl.load_workbook(f"{dir_name}{file_name}")
    sheet = book.active

    # Iterate over cells and set the number format
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = '0'
            elif isinstance(cell.value, float):
                cell.number_format = '0.0'
    # Set the number format for the specific columns
    for col in ["Температура воздуха(град)",
                "Температура точки росы(град)",
                "Давление на уровне вертолётной площадки(мм.рт.ст.)",
                "Давление на уровне моря(гПа)"]:
        column_index = df.columns.get_loc(col) + 1
        for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            for cell in row:
                cell.number_format = '0.0'

    # Установка высоты первой строки
    sheet.row_dimensions[1].height = 60
    # Создание объекта PatternFill для установки серого цвета
    grey_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    # Установка цвета фона для ячеек первой строки
    for cell in sheet[1]:
        cell.fill = grey_fill
    # Включение отображения фильтра в первой строке
    sheet.auto_filter.ref = sheet.dimensions

    # Включение переноса слов и выравнивание текста по центру
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    # Установка ширины столбцов
    for column in sheet.columns:
        column_letter = column[0].column_letter
        sheet.column_dimensions[column_letter].width = 14
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 7
    sheet.column_dimensions['F'].width = 16
    sheet.column_dimensions['G'].width = 16
    sheet.column_dimensions['H'].width = 18
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['P'].width = 21
    # Зафризить первую строку
    sheet.freeze_panes = 'A2'
    # Установка формата "дата" для столбца
    for cell in sheet['B']:
        cell.number_format = 'dd/mm/yyyy'
    # Save the Excel file
    book.save(f"{dir_name}{file_name}")

    result = messagebox.askquestion(title='Данные',
                                    message=f"В папке {dir_name} создан отчёт {file_name}\nОткрыть?")
    if result == 'yes':
        os.startfile(f"{dir_name}{file_name}")

# create_db_luna()
# delete_all_data()
# read_csv()
# print(select_from_db_test())
